import openpyxl
import sys
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']

N = int(sys.argv[1])

bold_font = Font(bold=True)

for i in range(1, N+1):
    row_cell = sheet.cell(row=i+1, column=1)
    row_cell.font = bold_font
    row_cell.value = i
    column_cell = sheet.cell(row=1, column=i+1)
    column_cell.font = bold_font
    column_cell.value = i

for i in range(1, N+1):
    for j in range(1, N+1):
        sheet.cell(row=i+1, column=j+1, value=i*j)

wb.save("multiplicationTable.xlsx")
